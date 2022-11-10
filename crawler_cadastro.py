import openpyxl

workbook = openpyxl.Workbook() #Cria uma planilha 
#workbook = openpyxl.load_workbook('south2.xlsx') #Lê uma planilha
sheet = workbook.active
#sheet2 = workbook2.active #Ativa a segunda planilha
sheet['A1']='Nome'
sheet['B1']='Data de Nascimento'
sheet['C1']='CPF'

def cadastrar():
    var1=2
    nome='i'
    plan = input('Digite o nome para planilha: ')
    while nome != '':
        
        print('\nCadastro')
        
        nome = input('Nome do Cliente: ')
        if nome.replace('\n','').replace('\r','').replace(' ','') == '' : break
        f=False
        for s in nome:
         if s in ['0','1','2','3','4','5','6','7','8','9']:
            f=True
        if f:
          print('Nao pode conter numero! Por favor Preencha Novamente !')
          continue
        
        dt_nascimento = input('Data de Nascimento: ')
        if nome.replace('\n','').replace('\r','').replace(' ','') == '' : break
        f=False
        for s in dt_nascimento:
          if s in ['a','b','c','d','e','f','g','h','i','j','k','l','m','n','o','p','q','r','s','t'
          'u','v','y','w','z']:
            f=True
        if f:
          print('Nao pode conter numeros! Por favor Preencha Novamente !')
          continue

        cpf = input('CPF: ')
        if cpf.replace('\n','').replace('\r','').replace(' ','') == '' : break
        f=False
        for s in cpf:
          if s in ['a','b','c','d','e','f','g','h','i','j','k','l','m','n','o','p','q','r','s','t'
          'u','v','y','w','z']:
            f=True
        if f:
          print('Nao pode conter numeros! Por favor Preencha Novamente !')
          continue

        sheet['A'+str(var1)] = nome
        sheet['B'+str(var1)] = dt_nascimento
        sheet['C'+str(var1)] = cpf
        
        var1+=1

    workbook.save(f'{plan}.xlsx')
    
def pesquisar():
    pesquisa = input('Digite o nome da planilha: ')
    workbook = openpyxl.load_workbook(f'{pesquisa}.xlsx')
    sheet2 = workbook.active #Ativa a segunda planilha
    print(sheet2)

def menu():
  escolha = input('1 para Cadastrar\n2 Selecionar Planilha\nSua Opção:')
  if(escolha == '1'):
    cadastrar()
  if(escolha == '2'):
    pesquisar()

menu()